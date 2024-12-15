from datetime import date, timezone, timedelta, datetime
from django.db.models import Count
from openpyxl.styles import Alignment, Side, Border

from robots.models import Robot
from openpyxl import Workbook

start_date = datetime.now() - timedelta(days=7)
end_date = datetime.now()

thins = Side(border_style="thin", color="000000")


def robots_report():

    wb = Workbook()
    data = (
        Robot.objects.values_list('model', 'version').
        annotate(robot_count=Count('*'))
        .filter(created__range=[start_date, end_date])
        .order_by('model')
    )
    # <QuerySet [('R2', 'D2', 12), ('R2', 'D3', 2), ('R2', 'D4', 15), ('R3', 'D4', 26)]>

    row = 3
    current_model = ''
    for tup in data:
        if tup[0] != current_model:
            row = 3
            current_model = tup[0]
            ws = wb.create_sheet(current_model)
            ws[f'B{row}'] = tup[0]
            ws[f'B{row}'].border = Border(top=thins, bottom=thins, left=thins, right=thins)
            ws[f'B{row}'].alignment = Alignment(horizontal='center')
            ws[f'C{row}'] = tup[1]
            ws[f'C{row}'].border = Border(top=thins, bottom=thins, left=thins, right=thins)
            ws[f'C{row}'].alignment = Alignment(horizontal='center')
            ws[f'D{row}'] = tup[2]
            ws[f'D{row}'].border = Border(top=thins, bottom=thins, left=thins, right=thins)
            ws[f'D{row}'].alignment = Alignment(horizontal='center')

            row += 1
        else:
            ws = wb[current_model]
            ws[f'B{row}'] = tup[0]
            ws[f'B{row}'].border = Border(top=thins, bottom=thins, left=thins, right=thins)
            ws[f'B{row}'].alignment = Alignment(horizontal='center')
            ws[f'C{row}'] = tup[1]
            ws[f'C{row}'].border = Border(top=thins, bottom=thins, left=thins, right=thins)
            ws[f'C{row}'].alignment = Alignment(horizontal='center')
            ws[f'D{row}'] = tup[2]
            ws[f'D{row}'].border = Border(top=thins, bottom=thins, left=thins, right=thins)
            ws[f'D{row}'].alignment = Alignment(horizontal='center')

            row += 1

    wb.remove(wb['Sheet'])

    wb.save('report.xlsx')






